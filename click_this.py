#-*- coding:utf-8 -*
import os, xlrd
from pyExcelerator import *
from openpyxl.reader.excel import load_workbook

INPUT_DIR = os.getcwd() + '\\input'

def process03(filename):
    workbook = xlrd.open_workbook(filename)
    '''with xlrd.open_workbook(filename) as workbook:'''
    sheet = workbook.sheet_by_index(0)
    nrows = sheet.nrows
    title_index = name_index = 0
    row_list = []
    for i in range(nrows):
        cur_row = sheet.row_values(i)
        if u"姓名" not in cur_row:
            title_index += 1
        else:
            name_index = cur_row.index(u"姓名")
            break        
    title = sheet.row_values(title_index)
    for j in range(title_index + 1,nrows):
        if sheet.row_values(j)[name_index] != '':
            row_list.append(sheet.row_values(j))
    return title,row_list

def process07(filename):
    row_list = []
    title = []
    title_index = name_index = 0
    try:
        workbook = load_workbook(filename)
        sheetnames = workbook.get_sheet_names()
        sheet = workbook.get_sheet_by_name(sheetnames[0])
        nrows = sheet.get_highest_row()
        ncols = sheet.get_highest_column()
        for r in range(nrows):
            for c in range(ncols):
                if sheet.cell(row = r,column = c).value == u"姓名":
                    title_index = r
                    name_index = c
                    break
            break
        for i in range(ncols):
            title.append(sheet.cell(row = title_index,column = i).value)
        for x in range(title_index + 1,nrows):
            temp = []
            for y in range(ncols):
                v = sheet.cell(row = x,column = y).value
                if (y == name_index)&(v == ''):
                    continue
                else:
                    temp.append(v)
            row_list.append(temp)
    except:
        print 'something wrong with' + filename
    return title,row_list

def main():
    title = []
    final_list = []
    for root,dirs,files in os.walk(INPUT_DIR):
            for f in files:
                    file_dir = os.path.join(INPUT_DIR,f)
                    print 'processing: ' + f
                    if file_dir.endswith('xls'):
                            title,row_list = process03(file_dir)
                    elif file_dir.endswith('xlsx'):
                            title,row_list = process07(file_dir)
                    else:
                            print f + '文件格式错误！'
                            continue
                    for line in row_list:
                            final_list.append(line)
    output = Workbook()
    new_sheet = output.add_sheet(u'汇总数据')
    col = 0
    for item in title:
            new_sheet.write(0,col,item)
            col += 1
    row = 1
    for line in final_list:
            col = 0
            for item in line:
                    if item is not None:
                            new_sheet.write(row,col,item)
                    else:
                            new_sheet.write(row,col,' ')
                    col += 1
            row += 1              
    output.save(u"汇总结果.xls")
    print 'OK!Your job is done.Please type ENTER and go to check it!'
   

if __name__ == '__main__':
    main()
    raw_input()
    
